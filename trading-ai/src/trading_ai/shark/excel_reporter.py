"""Daily Excel export for the universal trade journal (openpyxl)."""

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from trading_ai.governance.storage_architecture import shark_state_path
from trading_ai.shark.trade_journal import get_all_trades, get_summary_stats, get_trades_for_date

logger = logging.getLogger(__name__)

_FILL_WIN = PatternFill(fill_type="solid", fgColor="C6EFCE")
_FILL_LOSS = PatternFill(fill_type="solid", fgColor="FFC7CE")
_FILL_PENDING = PatternFill(fill_type="solid", fgColor="FFEB9C")


def _autosize_columns(ws: Any, max_width: int = 52) -> None:
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 10
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 1, max_width)


def generate_daily_excel(date_str: Optional[str] = None) -> Path:
    try:
        from zoneinfo import ZoneInfo

        et = ZoneInfo("America/New_York")
    except Exception:
        et = None  # type: ignore[assignment]
    if date_str is None:
        if et is not None:
            date_str = datetime.now(et).strftime("%Y-%m-%d")
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")

    trades = get_trades_for_date(date_str)
    stats = get_summary_stats(date_str)
    all_trades = get_all_trades()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.append(["Date", "Total Trades", "Wins", "Losses", "Win Rate", "Total P/L"])
    ws1.append(
        [
            date_str,
            stats.get("total_trades", 0),
            stats.get("wins", 0),
            stats.get("losses", 0),
            float(stats.get("win_rate", 0) or 0),
            float(stats.get("total_pnl", 0) or 0),
        ]
    )
    ws1.cell(2, 5).number_format = "0.00%"
    ws1.cell(2, 6).number_format = '"$"#,##0.00'

    ws2 = wb.create_sheet("All Trades Today")
    headers = [
        "Time",
        "Avenue",
        "Market",
        "Category",
        "Hunt Type",
        "Side",
        "Size",
        "Entry",
        "Exit",
        "P/L",
        "P/L %",
        "Outcome",
        "Claude",
        "Confidence",
        "Reasoning",
    ]
    ws2.append(headers)
    for t in sorted(trades, key=lambda x: str(x.get("timestamp") or "")):
        o = str(t.get("outcome", "pending")).lower()
        row = [
            str(t.get("timestamp", ""))[:19],
            t.get("avenue"),
            (t.get("question") or t.get("market_id") or "")[:500],
            t.get("category"),
            t.get("hunt_type"),
            t.get("side"),
            t.get("position_size_usd"),
            t.get("entry_price"),
            t.get("exit_price"),
            t.get("pnl_usd"),
            t.get("pnl_pct"),
            o,
            t.get("claude_decision"),
            t.get("claude_confidence"),
            (t.get("claude_reasoning") or "")[:800],
        ]
        ws2.append(row)
        fill = _FILL_PENDING
        if o == "win":
            fill = _FILL_WIN
        elif o == "loss":
            fill = _FILL_LOSS
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
    ws2.freeze_panes = "A2"
    _autosize_columns(ws2)

    ws3 = wb.create_sheet("By Avenue")
    ws3.append(["Avenue", "Trades", "Wins", "Losses", "Win Rate", "Total P/L", "Avg P/L per trade"])
    by_av = stats.get("by_avenue") or {}
    for av, b in sorted(by_av.items(), key=lambda x: str(x[0])):
        n = int(b.get("n", 0) or 0)
        w = int(b.get("wins", 0) or 0)
        l = int(b.get("losses", 0) or 0)
        pnl = float(b.get("pnl", 0) or 0)
        wr = (w / n) if n else 0.0
        avg = (pnl / n) if n else 0.0
        ws3.append([av, n, w, l, wr, pnl, avg])
    for r in range(2, ws3.max_row + 1):
        ws3.cell(r, 5).number_format = "0.00%"
        ws3.cell(r, 6).number_format = '"$"#,##0.00'
        ws3.cell(r, 7).number_format = '"$"#,##0.00'
    _autosize_columns(ws3)

    ws4 = wb.create_sheet("Claude Analysis")
    ws4.append(["Claude decision", "Closed trades", "Wins", "Win rate"])
    cl_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"n": 0, "w": 0})
    hunt_yes: Dict[str, int] = defaultdict(int)
    for t in trades:
        dec = str(t.get("claude_decision") or "").upper() or "(none)"
        o = str(t.get("outcome", "pending")).lower()
        if o == "pending":
            continue
        win = o == "win"
        cl_stats[dec]["n"] += 1
        if win:
            cl_stats[dec]["w"] += 1
        if dec == "YES":
            ht = str(t.get("hunt_type") or "unknown")
            hunt_yes[ht] += 1
    for dec, b in sorted(cl_stats.items(), key=lambda x: str(x[0])):
        n = int(b["n"])
        w = int(b["w"])
        ws4.append([dec, n, w, (w / n) if n else 0.0])
    ws4.append([])
    ws4.append(["Hunt types (Claude YES trades)", "Count"])
    for ht, c in sorted(hunt_yes.items(), key=lambda x: -x[1]):
        ws4.append([ht, c])
    _autosize_columns(ws4)

    ws5 = wb.create_sheet("Running Totals")
    starting = 0.0
    current = 0.0
    try:
        from trading_ai.shark.state_store import load_capital

        rec = load_capital()
        starting = float(rec.starting_capital or 0)
        current = float(rec.current_capital or 0)
    except Exception:
        pass
    closed_all = [t for t in all_trades if str(t.get("outcome", "pending")).lower() != "pending"]
    total_pnl = sum(float(t.get("pnl_usd", 0) or 0) for t in closed_all)
    dates = sorted({str(t.get("date")) for t in all_trades if t.get("date")})
    days_trading = len(dates)
    day_pnls: Dict[str, float] = defaultdict(float)
    for t in closed_all:
        d = str(t.get("date") or "")
        day_pnls[d] += float(t.get("pnl_usd", 0) or 0)
    best_day: Tuple[str, float] = ("", 0.0)
    worst_day: Tuple[str, float] = ("", 0.0)
    if day_pnls:
        best_day = ("", -1e18)
        worst_day = ("", 1e18)
        for d, v in day_pnls.items():
            if v > best_day[1]:
                best_day = (d, v)
            if v < worst_day[1]:
                worst_day = (d, v)
    streak = 0
    for d in reversed(dates):
        p = day_pnls.get(d, 0.0)
        if p > 1e-9:
            if streak >= 0:
                streak += 1
            else:
                break
        elif p < -1e-9:
            if streak <= 0:
                streak -= 1
            else:
                break
        else:
            break

    ws5.append(["Starting capital", "Current capital", "Total P/L (journal)", "Days trading"])
    ws5.append([starting, current, total_pnl, days_trading])
    ws5.append(["Best day", "Best day P/L", "Worst day", "Worst day P/L", "Current streak (days)"])
    ws5.append(
        [
            best_day[0],
            best_day[1],
            worst_day[0],
            worst_day[1],
            streak,
        ]
    )
    for c in (1, 2, 3):
        ws5.cell(2, c).number_format = '"$"#,##0.00'
    for c in (2, 4):
        ws5.cell(5, c).number_format = '"$"#,##0.00'
    _autosize_columns(ws5)

    for ws in (ws1, ws5):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws1[1][0].font = Font(bold=True)
    ws2[1][0].font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws2.cell(1, c).font = Font(bold=True)

    filename = f"ezras_trading_{date_str}.xlsx"
    path = shark_state_path(filename)
    wb.save(path)
    logger.info("Excel report saved: %s", path)
    return path
