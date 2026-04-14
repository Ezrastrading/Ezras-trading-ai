"""Universal trade journal + daily Excel (5 tests)."""

from __future__ import annotations

import uuid
import pytest


@pytest.fixture
def runtime_root(tmp_path, monkeypatch):
    monkeypatch.setenv("EZRAS_RUNTIME_ROOT", str(tmp_path))
    (tmp_path / "shark" / "state").mkdir(parents=True, exist_ok=True)
    return tmp_path


def test_log_trade_opened_returns_trade_id(runtime_root):
    from trading_ai.shark.models import ConfirmationResult, ExecutionIntent, HuntType
    from trading_ai.shark.trade_journal import get_all_trades, log_trade_opened

    intent = ExecutionIntent(
        market_id="KX-TEST",
        outlet="kalshi",
        side="yes",
        stake_fraction_of_capital=0.05,
        edge_after_fees=0.03,
        estimated_win_probability=0.55,
        hunt_types=[HuntType.NEAR_RESOLUTION],
        source="test",
        notional_usd=25.0,
        expected_price=0.5,
        shares=10,
    )
    conf = ConfirmationResult(
        actual_fill_price=0.48,
        actual_fill_size=10.0,
        slippage_pct=0.02,
        confirmed=True,
    )
    tid = log_trade_opened(intent, None, conf=conf, scored=None, execution_time_ms=42)
    assert len(tid) == 36
    uuid.UUID(tid)  # raises if invalid
    rows = get_all_trades()
    assert len(rows) == 1
    assert rows[0]["trade_id"] == tid
    assert rows[0]["outcome"] == "pending"
    assert rows[0]["execution_time_ms"] == 42


def test_log_trade_resolved_updates_entry(runtime_root):
    from trading_ai.shark.models import ConfirmationResult, ExecutionIntent, HuntType
    from trading_ai.shark.trade_journal import get_all_trades, log_trade_opened, log_trade_resolved

    intent = ExecutionIntent(
        market_id="poly:test",
        outlet="polymarket",
        side="no",
        stake_fraction_of_capital=0.05,
        edge_after_fees=0.02,
        estimated_win_probability=0.5,
        hunt_types=[HuntType.STRUCTURAL_ARBITRAGE],
        source="test",
        notional_usd=40.0,
        expected_price=0.4,
        shares=5,
    )
    conf = ConfirmationResult(0.4, 5.0, 0.0, True)
    tid = log_trade_opened(intent, None, conf=conf, scored=None)
    log_trade_resolved(tid, exit_price=1.0, pnl_usd=-12.5, outcome="loss")
    row = get_all_trades()[0]
    assert row["outcome"] == "loss"
    assert row["exit_price"] == 1.0
    assert row["pnl_usd"] == -12.5
    assert row["resolved_at"] is not None


def test_get_summary_stats_correct(runtime_root):
    from trading_ai.shark.trade_journal import get_summary_stats, log_sports_trade

    log_sports_trade("fanduel", "pick-a", -110, 10.0, "win", 8.5)
    log_sports_trade("draftkings", "pick-b", -110, 10.0, "loss", -10.0)
    s = get_summary_stats()
    assert s["total_trades"] == 2
    assert s["wins"] == 1
    assert s["losses"] == 1
    assert s["win_rate"] == pytest.approx(0.5)
    assert s["total_pnl"] == pytest.approx(-1.5, abs=0.01)


def test_generate_daily_excel_creates_file(runtime_root):
    from trading_ai.shark.excel_reporter import generate_daily_excel
    from trading_ai.shark.trade_journal import log_sports_trade

    log_sports_trade("fanduel", "evt-1", -110, 5.0, "win", 4.0)
    path = generate_daily_excel("2099-01-15")
    assert path.is_file()
    assert path.suffix == ".xlsx"
    assert "2099-01-15" in path.name


def test_generate_daily_excel_has_five_sheets(runtime_root):
    from openpyxl import load_workbook

    from trading_ai.shark.excel_reporter import generate_daily_excel

    path = generate_daily_excel("2099-06-01")
    wb = load_workbook(path)
    assert len(wb.sheetnames) == 5
    assert "Daily Summary" in wb.sheetnames
    assert "All Trades Today" in wb.sheetnames
    assert "By Avenue" in wb.sheetnames
    assert "Claude Analysis" in wb.sheetnames
    assert "Running Totals" in wb.sheetnames
